import sys, os, shutil
from openpyxl import load_workbook

class Room:
	def __init__(self, drow):
		self.name = drow['ROOM']
		self.rtype = drow['TYPE']
		self.cap = drow['CAP']
		self.discipline = drow['Discipline Priority'] or ''
	
	def __str__(self):
		return "%s|%s|Max %s" % (self.name, self.rtype, self.cap)

all_rooms = None

def readRoomFile(room_file, tmp_folder='~/Downloads'):
	global all_rooms
	all_rooms = {}
	
	# to prevent permissions problem we copy the file.
	temp_file = os.path.join(os.path.expanduser(tmp_folder), 'TempRooms.xlsx')
	shutil.copyfile(room_file, temp_file)
	
	wb = load_workbook(temp_file)
	sheet = wb.worksheets[0]
	header_row = [c.value for c in sheet[1]]
	for row in sheet.iter_rows(min_row=2, values_only=True):
		drow = dict(zip(header_row, row))
		room_name = drow['ROOM']
		all_rooms[room_name] = Room(drow)
	wb.close()

if __name__ == "__main__":
	if len(sys.argv) != 2:
		print('python3', sys.argv[0], 'Reference Folder')
		sys.exit(0)
	readRoomFile(sys.argv[1])
	print("\n".join(sorted(list(all_rooms.keys()))))
